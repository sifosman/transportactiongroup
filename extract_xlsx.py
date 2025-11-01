import openpyxl
from openpyxl import load_workbook
import csv

def extract_to_csv(file_path, output_path):
    wb = load_workbook(file_path)
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        with open(f"{output_path}_{sheet_name}.csv", 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)
            for row in sheet.iter_rows(values_only=True):
                writer.writerow(row)

if __name__ == "__main__":
    file_path = 'South Africa_N3_TCO_Model_ICEvEV.xlsx'
    extract_to_csv(file_path, 'South_Africa_N3_TCO_Model_ICEvEV')
    print("Extracted to CSV files")
